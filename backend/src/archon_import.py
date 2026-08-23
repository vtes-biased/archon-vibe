"""Archon Excel spreadsheet import: parses filled-in 'The Archon v1.5l' sheets
and imports results into an existing tournament, in place of NC approval.
"""

import io
import logging
from dataclasses import dataclass, field
from datetime import UTC, datetime

import msgspec
import openpyxl

from .db import (
    get_user_by_vekn_id,
    save_object,
    tournament_transaction,
)
from .models import (
    FinalsTable,
    ObjectType,
    PaymentStatus,
    Player,
    PlayerState,
    Score,
    Seat,
    Standing,
    Table,
    TableState,
    TournamentState,
    User,
)

logger = logging.getLogger(__name__)


@dataclass
class ArchonPlayer:
    number: int
    first_name: str
    last_name: str
    city: str
    vekn_id: str


@dataclass
class ArchonRoundTable:
    """A single table in a round: list of (player_number, vp) in seat order."""

    seats: list[tuple[int, float]]


@dataclass
class ArchonData:
    event_name: str
    num_rounds: int
    players: list[ArchonPlayer]
    rounds: list[list[ArchonRoundTable]]  # per round: list of tables
    finals: ArchonRoundTable | None = None


@dataclass
class ImportResult:
    success: bool
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    players_matched: int = 0
    rounds_imported: int = 0
    has_finals: bool = False


def _cell_float(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _cell_int(val) -> int:
    if val is None:
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0


def _cell_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def parse_archon_file(file_bytes: bytes) -> ArchonData:
    """Parse a filled-in Archon v1.5l spreadsheet. Sheet layout (0-indexed):
    1 tournament info, 2 methuselahs (players), 4-6 rounds 1-3, 7 final round."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets = wb.sheetnames

    ws_info = wb[sheets[1]]
    event_name = _cell_str(ws_info.cell(row=3, column=2).value) or "Imported Tournament"
    num_rounds = _cell_int(ws_info.cell(row=11, column=2).value) or 3

    ws_players = wb[sheets[2]]
    players: list[ArchonPlayer] = []
    row = 7
    while row <= ws_players.max_row:
        num_val = ws_players.cell(row=row, column=1).value
        if num_val is None:
            break
        player_num = _cell_int(num_val)
        if player_num == 0:
            break
        first_name = _cell_str(ws_players.cell(row=row, column=2).value)
        last_name = _cell_str(ws_players.cell(row=row, column=3).value)
        city = _cell_str(ws_players.cell(row=row, column=4).value)
        vekn_id = _cell_str(ws_players.cell(row=row, column=5).value)
        players.append(
            ArchonPlayer(
                number=player_num,
                first_name=first_name,
                last_name=last_name,
                city=city,
                vekn_id=vekn_id,
            )
        )
        row += 1

    rounds: list[list[ArchonRoundTable]] = []
    for round_idx in range(num_rounds):
        sheet_idx = 4 + round_idx
        if sheet_idx >= len(sheets):
            break
        ws_round = wb[sheets[sheet_idx]]
        tables: list[ArchonRoundTable] = []
        current_table_num = None
        current_seats: list[tuple[int, float]] = []
        row = 7
        while row <= ws_round.max_row:
            player_num_val = ws_round.cell(row=row, column=1).value
            if player_num_val is None:
                break
            player_num = _cell_int(player_num_val)
            if player_num == 0:
                break
            table_num = _cell_int(ws_round.cell(row=row, column=4).value)
            vp = _cell_float(ws_round.cell(row=row, column=5).value)

            if current_table_num is not None and table_num != current_table_num:
                tables.append(ArchonRoundTable(seats=current_seats))
                current_seats = []

            current_table_num = table_num
            current_seats.append((player_num, vp))
            row += 1

        if current_seats:
            tables.append(ArchonRoundTable(seats=current_seats))

        if tables:
            rounds.append(tables)

    finals: ArchonRoundTable | None = None
    if len(sheets) > 7:
        ws_finals = wb[sheets[7]]
        # Finals seating at rows ~14-18, col A = player number, col F = VPs
        finals_seats: list[tuple[int, float]] = []
        for row in range(14, 19):
            player_num_val = ws_finals.cell(row=row, column=1).value
            if player_num_val is None:
                continue
            player_num = _cell_int(player_num_val)
            if player_num == 0:
                continue
            vp = _cell_float(ws_finals.cell(row=row, column=6).value)
            finals_seats.append((player_num, vp))
        if finals_seats:
            finals = ArchonRoundTable(seats=finals_seats)

    wb.close()
    return ArchonData(
        event_name=event_name,
        num_rounds=num_rounds,
        players=players,
        rounds=rounds,
        finals=finals,
    )


def validate_archon_import(data: ArchonData, engine) -> list[str]:
    """Validate parsed archon data. Returns list of error strings (empty = valid)."""
    errors: list[str] = []

    players_with_vekn = [p for p in data.players if p.vekn_id]
    if len(players_with_vekn) < 4:
        errors.append(
            f"At least 4 players with VEKN IDs required, found {len(players_with_vekn)}"
        )

    if not data.rounds:
        errors.append("At least 1 round with data is required")

    vekn_ids = [p.vekn_id for p in data.players if p.vekn_id]
    seen: set[str] = set()
    for vid in vekn_ids:
        if vid in seen:
            errors.append(f"Duplicate VEKN ID: {vid}")
        seen.add(vid)

    for r_idx, tables in enumerate(data.rounds):
        for t_idx, table in enumerate(tables):
            table_size = len(table.seats)
            vps = [s[1] for s in table.seats]

            if table_size not in (4, 5):
                errors.append(
                    f"Round {r_idx + 1}, Table {t_idx + 1}: "
                    f"invalid table size {table_size} (must be 4 or 5)"
                )
                continue

            for s_idx, vp in enumerate(vps):
                if vp < 0 or vp > table_size:
                    errors.append(
                        f"Round {r_idx + 1}, Table {t_idx + 1}, Seat {s_idx + 1}: "
                        f"VP {vp} out of range [0, {table_size}]"
                    )
                if (vp * 2) != int(vp * 2):
                    errors.append(
                        f"Round {r_idx + 1}, Table {t_idx + 1}, Seat {s_idx + 1}: "
                        f"VP {vp} not in 0.5 increments"
                    )

            vp_error = engine.check_table_vps(vps)
            if vp_error:
                errors.append(
                    f"Round {r_idx + 1}, Table {t_idx + 1}: "
                    f"invalid VP distribution: {vp_error}"
                )

    if data.finals:
        fsize = len(data.finals.seats)
        if fsize not in (4, 5):
            errors.append(f"Finals: invalid table size {fsize} (must be 4 or 5)")
        else:
            fvps = [s[1] for s in data.finals.seats]
            vp_error = engine.check_table_vps(fvps)
            if vp_error:
                errors.append(f"Finals: invalid VP distribution: {vp_error}")

    return errors


async def apply_archon_import(
    tournament_uid: str,
    data: ArchonData,
    actor_uid: str,
    engine,
    broadcast_tournament_event=None,
    broadcast_user_event=None,
) -> ImportResult:
    """Apply parsed+validated archon data to an existing tournament.

    Builds tournament state directly (like vekn_tournament_sync.py) rather
    than replaying events.
    """
    user_by_number: dict[int, User] = {}
    unmatched: list[str] = []

    for p in data.players:
        if not p.vekn_id:
            unmatched.append(f"{p.first_name} {p.last_name} (no VEKN ID)")
            continue
        user = await get_user_by_vekn_id(p.vekn_id)
        if user:
            user_by_number[p.number] = user
        else:
            unmatched.append(f"{p.first_name} {p.last_name} (VEKN ID: {p.vekn_id})")

    if unmatched:
        return ImportResult(
            success=False,
            errors=[f"Unmatched players: {', '.join(unmatched)}"],
        )

    uid_by_number: dict[int, str] = {num: u.uid for num, u in user_by_number.items()}

    built_rounds: list[list[Table]] = []
    # player_scores is prelim+finals (feeds Player.result); player_prelim is
    # prelim-only (feeds standings) — finals must never fold into it.
    player_scores: dict[str, dict] = {}
    player_prelim: dict[str, dict] = {}
    for user_uid in uid_by_number.values():
        player_scores[user_uid] = {"gw": 0.0, "vp": 0.0, "tp": 0}
        player_prelim[user_uid] = {"gw": 0.0, "vp": 0.0, "tp": 0}

    for tables in data.rounds:
        round_tables: list[Table] = []
        for table in tables:
            vps = [s[1] for s in table.seats]
            table_size = len(vps)
            adjustments = [0.0] * table_size
            gws = engine.compute_gw(vps, adjustments)
            tps = engine.compute_tp(table_size, vps, adjustments)

            seats: list[Seat] = []
            for i, (player_num, vp) in enumerate(table.seats):
                user_uid = uid_by_number[player_num]
                gw = int(gws[i])
                tp = int(tps[i])
                seats.append(
                    Seat(
                        player_uid=user_uid,
                        result=Score(gw=gw, vp=vp, tp=tp),
                    )
                )
                player_scores[user_uid]["gw"] += gw
                player_scores[user_uid]["vp"] += vp
                player_scores[user_uid]["tp"] += tp
                player_prelim[user_uid]["gw"] += gw
                player_prelim[user_uid]["vp"] += vp
                player_prelim[user_uid]["tp"] += tp

            round_tables.append(
                Table(
                    seating=seats,
                    state=TableState.FINISHED,
                )
            )
        built_rounds.append(round_tables)

    built_finals: FinalsTable | None = None
    finals_winner_uid = ""
    if data.finals:
        fvps = [s[1] for s in data.finals.seats]
        fsize = len(fvps)
        f_adjustments = [0.0] * fsize
        # Must use compute_gw_finals (not prelim compute_gw): the engine's finals
        # refresh recomputes with this rule, so a mismatch would silently drift the GW.
        f_uids = [uid_by_number[pn] for pn, _ in data.finals.seats]
        f_gws = engine.compute_gw_finals(fvps, f_adjustments, f_uids, f_uids)
        f_tps = engine.compute_tp(fsize, fvps, f_adjustments)

        finals_seats: list[Seat] = []
        seed_order: list[str] = []
        for i, (player_num, vp) in enumerate(data.finals.seats):
            user_uid = uid_by_number[player_num]
            gw = int(f_gws[i])
            tp = int(f_tps[i])
            finals_seats.append(
                Seat(
                    player_uid=user_uid,
                    result=Score(gw=gw, vp=vp, tp=tp),
                )
            )
            seed_order.append(user_uid)
            player_scores[user_uid]["gw"] += gw
            player_scores[user_uid]["vp"] += vp
            player_scores[user_uid]["tp"] += tp
            if gw == 1:
                finals_winner_uid = user_uid

        built_finals = FinalsTable(
            seating=finals_seats,
            seed_order=seed_order,
            state=TableState.FINISHED,
        )

    finalist_uids = set()
    if data.finals:
        finalist_uids = {uid_by_number[s[0]] for s in data.finals.seats}

    players_list: list[Player] = []
    standings_list: list[Standing] = []
    for _player_num, user in user_by_number.items():
        user_uid = user.uid
        scores = player_scores[user_uid]
        players_list.append(
            Player(
                user_uid=user_uid,
                state=PlayerState.FINISHED,
                payment_status=PaymentStatus.PAID,
                toss=0,
                result=Score(
                    gw=int(scores["gw"]),
                    vp=scores["vp"],
                    tp=int(scores["tp"]),
                ),
                finalist=user_uid in finalist_uids,
            )
        )
        prelim = player_prelim[user_uid]
        standings_list.append(
            Standing(
                user_uid=user_uid,
                gw=float(prelim["gw"]),
                vp=prelim["vp"],
                tp=int(prelim["tp"]),
                toss=0,
                finalist=user_uid in finalist_uids,
            )
        )

    # The engine's own order, so an import cannot rank a sheet the engine never would.
    standings_list = msgspec.json.decode(
        engine.sort_standings(msgspec.json.encode(standings_list).decode()),
        type=list[Standing],
    )

    winner_uid = finals_winner_uid
    if not winner_uid and standings_list:
        winner_uid = standings_list[0].user_uid

    async with tournament_transaction(tournament_uid) as (tournament, tx_conn):
        if not tournament:
            return ImportResult(success=False, errors=["Tournament not found"])

        tournament.state = TournamentState.FINISHED
        tournament.players = players_list
        tournament.rounds = built_rounds
        tournament.finals = built_finals
        tournament.winner = winner_uid
        tournament.standings = standings_list
        tournament.modified = datetime.now(UTC)

        tournament_bd = await save_object(
            ObjectType.TOURNAMENT,
            tournament.uid,
            msgspec.to_builtins(tournament),
            conn=tx_conn,
        )

    # Post-effects run outside the transaction — must not hold the row locked
    # while pushing to VEKN or recomputing ratings.
    logger.info(
        f"Archon import completed for tournament {tournament_uid}: "
        f"{len(players_list)} players, {len(built_rounds)} rounds, "
        f"finals={'yes' if built_finals else 'no'}"
    )

    if broadcast_tournament_event:
        broadcast_tournament_event(tournament_bd)

    try:
        from .ratings import (
            rating_category_for_tournament,
            recompute_ratings_for_players,
            recompute_wins,
        )

        player_uids = {p.user_uid for p in players_list if p.user_uid}
        category = rating_category_for_tournament(tournament)
        results = await recompute_ratings_for_players(player_uids, category)
        results += await recompute_wins(player_uids)
        if broadcast_user_event:
            for _user, bd in results:
                broadcast_user_event(bd)
    except Exception:
        logger.exception(f"Error recomputing ratings for {tournament_uid}")

    try:
        from .routes.tournaments import _maybe_push_vekn, maybe_submit_twda

        await maybe_submit_twda(tournament)
        await _maybe_push_vekn(tournament)
    except Exception:
        logger.exception(f"Error in post-import effects for {tournament_uid}")

    return ImportResult(
        success=True,
        players_matched=len(players_list),
        rounds_imported=len(built_rounds),
        has_finals=built_finals is not None,
    )

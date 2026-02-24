#!/usr/bin/env python3
"""Compare our seating algorithm with optimal seatings from Excel."""

import os
import sys

# Add parent directory for openpyxl
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl


def get_table_sizes(n):
    """Get VEKN table sizes for n players: 4 or 5 player tables only."""
    if n < 4:
        return []
    # n = 5*a + 4*b where a, b >= 0, minimize tables (maximize 5s)
    for num_5s in range(n // 5, -1, -1):
        leftover = n - 5 * num_5s
        if leftover >= 0 and leftover % 4 == 0:
            num_4s = leftover // 4
            return [5] * num_5s + [4] * num_4s
    return None  # Not possible (6, 7, 11 - staggered)


def parse_round_cells(cells, n_players):
    """Parse a round from Excel cells. Use n_players to determine table splits.

    The Excel format lists players in sequence. We ignore any gaps and split
    purely based on VEKN table sizes (4 or 5 players per table).
    """
    # Get expected table sizes
    sizes = get_table_sizes(n_players)
    if not sizes:
        return None  # Staggered format, skip

    # Collect all player numbers (ignore None gaps - they're just formatting)
    players = [int(c) for c in cells if isinstance(c, (int, float))]

    if len(players) != n_players:
        return None  # Unexpected number of players

    # Split by VEKN table sizes
    tables = []
    idx = 0
    for size in sizes:
        tables.append(players[idx : idx + size])
        idx += size
    return tables


def extract_seatings_3r(sheet, max_players=50):
    """Extract 3-round seatings from Excel sheet."""
    seatings = {}
    row = 1

    while row < 1500:
        # Look for "N Players" pattern in column D
        d_val = sheet.cell(row=row, column=4).value
        e_val = sheet.cell(row=row, column=5).value

        if isinstance(d_val, (int, float)) and e_val and "Players" in str(e_val):
            n_players = int(d_val)
            if n_players > max_players:
                row += 1
                continue

            rounds = []
            # Read the next 3 rows for rounds
            for r_offset in range(1, 6):
                r = row + r_offset
                label = sheet.cell(row=r, column=5).value
                if label and "Round" in str(label):
                    cells = [sheet.cell(row=r, column=c).value for c in range(6, 35)]
                    tables = parse_round_cells(cells, n_players)
                    if tables:
                        rounds.append(tables)

            if rounds:
                seatings[n_players] = rounds

        row += 1

    return seatings


def compute_score(rounds):
    """Compute score for given rounds."""
    # Find max player number
    all_players = set()
    for r in rounds:
        for t in r:
            all_players.update(t)

    n = max(all_players)
    rounds_count = len(rounds)

    # Initialize measure matrix (n+1 x n+1 x 8) - 1-indexed
    measure = [[[0] * 8 for _ in range(n + 1)] for _ in range(n + 1)]

    # Position vectors
    POSITIONS_4 = [
        [1, 4, 1, 1, 0, 0, 0, 0],
        [1, 4, 2, 0, 1, 0, 0, 0],
        [1, 4, 3, 0, 0, 1, 0, 0],
        [1, 4, 4, 0, 0, 0, 1, 0],
    ]
    POSITIONS_5 = [
        [1, 5, 1, 1, 0, 0, 0, 0],
        [1, 5, 2, 0, 1, 0, 0, 0],
        [1, 5, 3, 0, 0, 1, 0, 0],
        [1, 5, 4, 0, 0, 0, 1, 0],
        [1, 5, 4, 0, 0, 0, 0, 1],
    ]

    # Opponent vectors
    OPPONENTS_4 = [
        [1, 1, 0, 0, 0, 0, 1, 0],  # prey
        [1, 0, 0, 0, 0, 1, 0, 1],  # cross
        [1, 0, 0, 0, 1, 0, 1, 0],  # predator
    ]
    OPPONENTS_5 = [
        [1, 1, 0, 0, 0, 0, 1, 0],  # prey
        [1, 0, 1, 0, 0, 0, 0, 1],  # grand-prey
        [1, 0, 0, 1, 0, 0, 0, 1],  # grand-predator
        [1, 0, 0, 0, 1, 0, 1, 0],  # predator
    ]

    for rnd in rounds:
        for table in rnd:
            size = len(table)
            if size < 4 or size > 5:
                continue
            positions = POSITIONS_4 if size == 4 else POSITIONS_5
            opponents = OPPONENTS_4 if size == 4 else OPPONENTS_5

            for seat, player in enumerate(table):
                # Position on diagonal
                for k in range(8):
                    measure[player][player][k] += positions[seat][k]

                # Opponent relationships
                for rel in range(size - 1):
                    opp_seat = (seat + rel + 1) % size
                    opp = table[opp_seat]
                    for k in range(8):
                        measure[player][opp][k] += opponents[rel][k]

    # Compute rule violations
    r1 = r2 = r4 = r5 = r6 = r7 = r9 = 0
    vps_list = []
    transfers_list = []

    for i in range(1, n + 1):
        played = measure[i][i][0]
        if played > 0:
            vps = measure[i][i][1] / played
            transfers = measure[i][i][2] / played
            vps_list.append(vps)
            transfers_list.append(transfers)

            # R5, R7: seat violations
            for seat in range(3, 8):
                if measure[i][i][seat] > 1:
                    r7 += 1
                    if seat == 7:  # fifth seat
                        r5 += 1

    for i in range(1, n + 1):
        for j in range(1, i):
            opp_count = measure[i][j][0]
            if opp_count > 1:
                r4 += 1
                if opp_count >= rounds_count:
                    r2 += 1

                # R6: same relative position (prey, grand-prey, grand-pred, pred, cross)
                for k in range(1, 6):
                    if measure[i][j][k] > 1:
                        r6 += 1
                        if k == 1 or k == 4:  # prey or predator
                            r1 += 1

                # R9: position group (neighbour, non-neighbour)
                for k in range(6, 8):
                    if measure[i][j][k] > 1:
                        r9 += 1

    # R3, R8: standard deviations
    if vps_list:
        mean_vps = sum(vps_list) / len(vps_list)
        r3 = (sum((v - mean_vps) ** 2 for v in vps_list) / len(vps_list)) ** 0.5
        mean_transfers = sum(transfers_list) / len(transfers_list)
        r8 = (
            sum((t - mean_transfers) ** 2 for t in transfers_list) / len(transfers_list)
        ) ** 0.5
    else:
        r3 = r8 = 0

    return [r1, r2, r3, r4, r5, r6, r7, r8, r9]


def run_rust_and_get_score(n_players, rounds_count=3):
    """Run Rust seating and get score by calling cargo test."""
    # TODO: implement Rust benchmark parsing
    return None


def main():
    print("Loading Excel file...")
    wb = openpyxl.load_workbook("../INCOMING/thearchon1.5l.xlsx", data_only=True)

    sheet = wb["Optimal Seating 3R+F"]
    print("\n=== Optimal Seating 3R+F ===")

    seatings = extract_seatings_3r(sheet, max_players=50)
    print(
        f"Found optimal seatings for {len(seatings)} player counts: {sorted(seatings.keys())}"
    )

    print("\nExcel optimal scores (3 rounds):")
    print("-" * 80)

    results = []
    for n_players in sorted(seatings.keys()):
        rounds = seatings[n_players]

        # For staggered counts (6, 7, 11), they might have different round counts
        if n_players in [6, 7, 11]:
            continue  # Skip staggered for now

        # Take first 3 rounds
        if len(rounds) < 3:
            continue

        rounds_3 = rounds[:3]
        score = compute_score(rounds_3)

        hard = int(score[0] + score[1])  # R1 + R2
        status = "✓" if hard == 0 else "✗"

        results.append((n_players, score))
        print(
            f"{n_players:3d}p: {status} R1={score[0]:.0f} R2={score[1]:.0f} R3={score[2]:.2f} R4={score[3]:.0f} R5={score[4]:.0f} R6={score[5]:.0f} R7={score[6]:.0f} R8={score[7]:.2f} R9={score[8]:.0f}"
        )

    # Now let's compare with our Rust implementation
    # We'll output a simple comparison format
    print("\n" + "=" * 80)
    print("To compare with Rust, run benchmarks and compare R1-R9 values")
    print("=" * 80)

    # Output data for Rust comparison
    print("\nExcel optimal data (for Rust comparison test):")
    for n_players, score in results[:15]:  # First 15 for focused comparison
        print(f"({n_players}, {score}),")


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    main()
